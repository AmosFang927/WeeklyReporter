"""
Data-Output-Agent: 输出处理器
处理所有数据输出格式：JSON、Excel、飞书上传、邮件发送
"""

import json
import os
import asyncio
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 导入邮件和飞书相关模块
import sys
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from config import (
    OUTPUT_DIR, 
    EMAIL_SENDER, 
    EMAIL_PASSWORD,
    SMTP_SERVER,
    SMTP_PORT,
    FEISHU_APP_ID,
    FEISHU_APP_SECRET,
    FEISHU_PARENT_NODE
)
from modules.email_sender import EmailSender
from modules.feishu_uploader import FeishuUploader


class OutputProcessor:
    """数据输出处理器"""
    
    def __init__(self):
        self.output_dir = Path(OUTPUT_DIR)
        self.output_dir.mkdir(exist_ok=True)
        
        # 初始化邮件和飞书上传器
        self.email_sender = EmailSender()
        self.feishu_uploader = FeishuUploader()
    
    async def generate_json(self, data: List[Dict[str, Any]], partner_name: str, 
                          start_date: str, end_date: str) -> Dict[str, Any]:
        """生成JSON文件"""
        try:
            # 准备JSON数据
            json_data = {
                'partner_name': partner_name,
                'date_range': {
                    'start_date': start_date,
                    'end_date': end_date
                },
                'generated_at': datetime.now().isoformat(),
                'total_records': len(data),
                'total_amount': sum(record.get('commission_amount', 0) for record in data),
                'conversions': data
            }
            
            # 生成文件名
            filename = f"{partner_name}_ConversionReport_{start_date}_to_{end_date}.json"
            file_path = self.output_dir / filename
            
            # 写入JSON文件
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
            
            return {
                'success': True,
                'message': f'JSON文件生成成功',
                'file_path': str(file_path),
                'total_records': len(data),
                'total_amount': json_data['total_amount']
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'JSON文件生成失败: {str(e)}'
            }
    
    async def generate_excel(self, data: List[Dict[str, Any]], partner_name: str, 
                           start_date: str, end_date: str) -> Dict[str, Any]:
        """生成Excel文件"""
        try:
            # 转换为DataFrame
            df = pd.DataFrame(data)
            
            if df.empty:
                return {
                    'success': False,
                    'error': 'No data to generate Excel file'
                }
            
            # 生成文件名
            filename = f"{partner_name}_ConversionReport_{start_date}_to_{end_date}.xlsx"
            file_path = self.output_dir / filename
            
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Conversion Report"
            
            # 添加标题行
            title_row = [
                'Transaction ID', 'Partner Name', 'Offer Name', 'Sale Amount', 
                'Currency', 'Commission Amount', 'Conversion Date', 'Status'
            ]
            ws.append(title_row)
            
            # 设置标题行样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, cell in enumerate(ws[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
            
            # 添加数据行
            for record in data:
                row = [
                    record.get('transaction_id', ''),
                    record.get('partner_name', ''),
                    record.get('offer_name', ''),
                    record.get('sale_amount', 0),
                    record.get('currency', 'USD'),
                    record.get('commission_amount', 0),
                    record.get('conversion_date', ''),
                    record.get('status', '')
                ]
                ws.append(row)
            
            # 添加合计行
            total_amount = sum(record.get('commission_amount', 0) for record in data)
            total_row = ['', '', '', '', '', f'Total: ${total_amount:,.2f}', '', '']
            ws.append(total_row)
            
            # 设置合计行样式
            total_cell = ws[f'F{ws.max_row}']
            total_cell.font = Font(bold=True)
            total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # 保存文件
            wb.save(file_path)
            
            return {
                'success': True,
                'message': f'Excel文件生成成功',
                'file_path': str(file_path),
                'total_records': len(data),
                'total_amount': total_amount
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Excel文件生成失败: {str(e)}'
            }
    
    async def upload_to_feishu(self, data: List[Dict[str, Any]], partner_name: str, 
                             start_date: str, end_date: str) -> Dict[str, Any]:
        """上传到飞书"""
        try:
            # 先生成Excel文件
            excel_result = await self.generate_excel(data, partner_name, start_date, end_date)
            
            if not excel_result['success']:
                return {
                    'success': False,
                    'error': f'生成Excel文件失败: {excel_result["error"]}'
                }
            
            # 上传到飞书
            upload_result = await self.feishu_uploader.upload_file(
                file_path=excel_result['file_path'],
                folder_name=f"{partner_name}合作伙伴",
                file_name=f"{partner_name}_ConversionReport_{start_date}_to_{end_date}.xlsx"
            )
            
            if upload_result['success']:
                return {
                    'success': True,
                    'message': '飞书上传成功',
                    'feishu_url': upload_result.get('file_url'),
                    'total_records': len(data),
                    'total_amount': sum(record.get('commission_amount', 0) for record in data)
                }
            else:
                return {
                    'success': False,
                    'error': f'飞书上传失败: {upload_result["error"]}'
                }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'飞书上传失败: {str(e)}'
            }
    
    async def send_email(self, data: List[Dict[str, Any]], partner_name: str, 
                        start_date: str, end_date: str, recipients: List[str]) -> Dict[str, Any]:
        """发送邮件"""
        try:
            # 先生成Excel文件
            excel_result = await self.generate_excel(data, partner_name, start_date, end_date)
            
            if not excel_result['success']:
                return {
                    'success': False,
                    'error': f'生成Excel文件失败: {excel_result["error"]}'
                }
            
            # 准备邮件内容
            total_amount = sum(record.get('commission_amount', 0) for record in data)
            
            subject = f"{partner_name} Partner - 转化数据报告 ({start_date} to {end_date})"
            
            body = f"""
            <html>
            <body>
                <h2>{partner_name} Partner 转化数据报告</h2>
                <p><strong>报告日期范围:</strong> {start_date} 到 {end_date}</p>
                <p><strong>总记录数:</strong> {len(data)} 条</p>
                <p><strong>总金额:</strong> ${total_amount:,.2f}</p>
                <hr>
                <p>详细数据请查看附件Excel文件。</p>
                <p>如有疑问，请联系我们。</p>
                <br>
                <p>Best regards,<br>ByteC Network Team</p>
            </body>
            </html>
            """
            
            # 发送邮件
            send_result = await self.email_sender.send_email(
                recipients=recipients,
                subject=subject,
                body=body,
                attachments=[excel_result['file_path']],
                is_html=True
            )
            
            if send_result['success']:
                return {
                    'success': True,
                    'message': '邮件发送成功',
                    'sent_to': recipients,
                    'total_records': len(data),
                    'total_amount': total_amount
                }
            else:
                return {
                    'success': False,
                    'error': f'邮件发送失败: {send_result["error"]}'
                }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'邮件发送失败: {str(e)}'
            }
    
    async def process_all_formats(self, data: List[Dict[str, Any]], partner_name: str, 
                                start_date: str, end_date: str, 
                                recipients: Optional[List[str]] = None,
                                skip_feishu: bool = False,
                                skip_email: bool = False) -> Dict[str, Any]:
        """处理所有输出格式"""
        results = {}
        
        # 生成JSON
        json_result = await self.generate_json(data, partner_name, start_date, end_date)
        results['json'] = json_result
        
        # 生成Excel
        excel_result = await self.generate_excel(data, partner_name, start_date, end_date)
        results['excel'] = excel_result
        
        # 上传飞书
        if not skip_feishu:
            feishu_result = await self.upload_to_feishu(data, partner_name, start_date, end_date)
            results['feishu'] = feishu_result
        
        # 发送邮件
        if not skip_email and recipients:
            email_result = await self.send_email(data, partner_name, start_date, end_date, recipients)
            results['email'] = email_result
        
        # 汇总结果
        success_count = sum(1 for result in results.values() if result.get('success'))
        total_count = len(results)
        
        return {
            'success': success_count == total_count,
            'results': results,
            'summary': {
                'total_records': len(data),
                'total_amount': sum(record.get('commission_amount', 0) for record in data),
                'processed_formats': list(results.keys()),
                'success_count': success_count,
                'total_count': total_count
            }
        } 