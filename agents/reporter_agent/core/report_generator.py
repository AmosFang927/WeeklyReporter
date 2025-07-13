#!/usr/bin/env python3
"""
报表生成器
从数据库生成Excel报表，并发送到飞书和邮件
"""

import os
import sys
import asyncio
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 导入ByteC-Network-Agent的现有模块
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from modules.feishu_uploader import FeishuUploader
from modules.email_sender import EmailSender
from .database import PostbackDatabase, PartnerSummary

logger = logging.getLogger(__name__)

class ReportGenerator:
    """报表生成器"""
    
    def __init__(self, output_dir: str = "output", 
                 global_email_disabled: bool = False):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # 初始化数据库连接
        self.db = PostbackDatabase()
        
        # 初始化飞书上传和邮件发送器
        self.feishu_uploader = FeishuUploader()
        self.email_sender = EmailSender(global_email_disabled=global_email_disabled)
        
        # 邮件配置
        self.partner_email_mapping = {
            'InvolveAsia': ['partners@involveasia.com'],
            'Rector': ['rector@partners.com'],
            'DeepLeaper': ['deepleaper@partners.com'],
            'ByteC': ['bytec@partners.com'],
            'RAMPUP': ['rampup@partners.com'],
            'ALL': ['AmosFang927@gmail.com']
        }
        
        self.partner_email_enabled = {
            'InvolveAsia': True,
            'Rector': True,
            'DeepLeaper': True,
            'ByteC': True,
            'RAMPUP': True,
            'ALL': True
        }
    
    async def generate_partner_report(self, partner_name: str = "ALL",
                                    start_date: datetime = None,
                                    end_date: datetime = None,
                                    send_email: bool = True,
                                    upload_feishu: bool = True,
                                    self_email: bool = False) -> Dict[str, Any]:
        """
        生成Partner报表
        
        Args:
            partner_name: Partner名称 (ALL, InvolveAsia, Rector, DeepLeaper, ByteC, RAMPUP)
            start_date: 开始日期
            end_date: 结束日期
            send_email: 是否发送邮件
            upload_feishu: 是否上传到飞书
            
        Returns:
            Dict[str, Any]: 报表生成结果
        """
        try:
            logger.info(f"🚀 开始生成 {partner_name} 报表")
            
            # 设置默认日期范围（过去7天）
            if not end_date:
                end_date = datetime.now()
            if not start_date:
                start_date = end_date - timedelta(days=7)
            
            # 获取数据
            logger.info(f"🔍 查询数据: {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}")
            
            # 获取转化数据
            df = await self.db.get_conversion_dataframe(partner_name, start_date, end_date)
            
            if df.empty:
                logger.warning(f"⚠️ 没有找到 {partner_name} 的转化数据")
                print(f"📋 提示: {partner_name} 在 {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')} 期间没有转化数据，將生成空報表並繼續執行")
                
                # 创建空的partner_summaries
                partner_summaries = []
            else:
                # 获取Partner汇总
                partner_summaries = await self.db.get_partner_summary(partner_name, start_date, end_date)
            
            # 生成Excel文件
            excel_files = await self._generate_excel_files(df, partner_summaries, partner_name, start_date, end_date)
            
            # 計算總金額（處理空數據情況）
            if df.empty or 'Sale Amount (USD)' not in df.columns:
                total_amount = 0.0
            else:
                total_amount = df['Sale Amount (USD)'].sum()
            
            result = {
                'success': True,
                'partner_name': partner_name,
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'total_records': len(df),
                'total_amount': total_amount,
                'excel_files': excel_files,
                'partner_summaries': [summary.to_dict() for summary in partner_summaries]
            }
            
            # 飞书上传
            if upload_feishu and excel_files:
                logger.info("📤 开始上传到飞书")
                feishu_result = await self._upload_to_feishu(excel_files)
                result['feishu_upload'] = feishu_result
            
            # 邮件发送
            if send_email and excel_files:
                logger.info("✉️ 开始发送邮件")
                email_result = await self._send_emails(partner_summaries, excel_files, start_date, end_date, self_email)
                result['email_result'] = email_result
            
            logger.info(f"✅ {partner_name} 报表生成完成")
            return result
            
        except Exception as e:
            logger.error(f"❌ 生成报表失败: {e}")
            return {
                'success': False,
                'error': str(e),
                'partner_name': partner_name
            }
    
    async def _generate_excel_files(self, df: pd.DataFrame, 
                                  partner_summaries: List[PartnerSummary],
                                  partner_name: str, 
                                  start_date: datetime,
                                  end_date: datetime) -> List[str]:
        """生成Excel文件"""
        excel_files = []
        
        try:
            if partner_name.upper() == 'ALL':
                # 为每个Partner生成单独的文件
                for summary in partner_summaries:
                    partner_df = df[df['Partner'] == summary.partner_name]
                    if not partner_df.empty:
                        file_path = await self._create_excel_file(
                            partner_df, 
                            summary.partner_name, 
                            start_date, 
                            end_date
                        )
                        excel_files.append(file_path)
                        
                        # 为每个Partner的汇总添加文件路径
                        summary.file_path = file_path
                
                # 生成总汇总文件
                main_file = await self._create_excel_file(df, "AllPartners", start_date, end_date)
                excel_files.insert(0, main_file)  # 插入到第一个位置
                
            else:
                # 生成单个Partner的文件（即使是空數據也要生成）
                file_path = await self._create_excel_file(df, partner_name, start_date, end_date)
                excel_files.append(file_path)
                
                # 为汇总添加文件路径
                if partner_summaries:
                    partner_summaries[0].file_path = file_path
            
            logger.info(f"✅ 成功生成 {len(excel_files)} 个Excel文件")
            return excel_files
            
        except Exception as e:
            logger.error(f"❌ 生成Excel文件失败: {e}")
            raise
    
    async def _create_excel_file(self, df: pd.DataFrame, 
                               partner_name: str,
                               start_date: datetime,
                               end_date: datetime) -> str:
        """创建单个Excel文件"""
        try:
            # 生成文件名
            date_str = f"{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}"
            filename = f"{partner_name}_ConversionReport_{date_str}.xlsx"
            file_path = self.output_dir / filename
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = f"{partner_name} Conversions"
            
            # 设置标题样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border
                    
                    # 设置标题行样式
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # 格式化金额列
                    if c_idx == 7 and r_idx > 1:  # Sale Amount (USD) 列
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 8 and r_idx > 1:  # Payout (USD) 列
                        cell.number_format = '"$"#,##0.00'
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            # 添加汇总信息
            last_row = len(df) + 3
            ws.cell(row=last_row, column=1, value="汇总信息:").font = Font(bold=True)
            ws.cell(row=last_row + 1, column=1, value=f"Partner: {partner_name}")
            ws.cell(row=last_row + 2, column=1, value=f"日期范围: {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}")
            ws.cell(row=last_row + 3, column=1, value=f"总记录数: {len(df):,}")
            
            # 處理空數據情況
            if df.empty or 'Sale Amount (USD)' not in df.columns:
                ws.cell(row=last_row + 4, column=1, value="总金额: $0.00")
            else:
                ws.cell(row=last_row + 4, column=1, value=f"总金额: ${df['Sale Amount (USD)'].sum():,.2f}")
            
            ws.cell(row=last_row + 5, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            # 保存文件
            wb.save(file_path)
            logger.info(f"✅ 成功创建Excel文件: {filename}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"❌ 创建Excel文件失败: {e}")
            raise
    
    async def _upload_to_feishu(self, excel_files: List[str]) -> Dict[str, Any]:
        """上传到飞书"""
        try:
            # 使用原有的FeishuUploader
            upload_result = self.feishu_uploader.upload_files(excel_files)
            
            if upload_result['success']:
                logger.info(f"✅ 飞书上传成功: {upload_result['success_count']} 个文件")
            else:
                logger.warning(f"⚠️ 飞书上传部分失败: 成功 {upload_result['success_count']} 个，失败 {upload_result['failed_count']} 个")
            
            return upload_result
            
        except Exception as e:
            logger.error(f"❌ 飞书上传失败: {e}")
            return {
                'success': False,
                'error': str(e),
                'success_count': 0,
                'failed_count': len(excel_files)
            }
    
    async def _send_emails(self, partner_summaries: List[PartnerSummary],
                         excel_files: List[str],
                         start_date: datetime,
                         end_date: datetime,
                         self_email: bool = False) -> Dict[str, Any]:
        """发送邮件"""
        try:
            # 准备Partner邮件数据
            partner_summary_dict = {}
            for summary in partner_summaries:
                partner_summary_dict[summary.partner_name] = {
                    'records': summary.total_records,
                    'amount_formatted': summary.amount_formatted,
                    'file_path': getattr(summary, 'file_path', None),
                    'sources': summary.sources,
                    'sources_count': summary.sources_count
                }
            
            # 使用原有的EmailSender发送邮件
            email_result = self.email_sender.send_partner_reports(
                partner_summary_dict,
                None,  # feishu_upload_result
                end_date,
                start_date,
                self_email
            )
            
            if email_result['success']:
                logger.info(f"✅ 邮件发送成功: {email_result['total_sent']} 封")
            else:
                logger.warning(f"⚠️ 邮件发送部分失败: 成功 {email_result['total_sent']} 封，失败 {email_result['total_failed']} 封")
            
            return email_result
            
        except Exception as e:
            logger.error(f"❌ 邮件发送失败: {e}")
            return {
                'success': False,
                'error': str(e),
                'total_sent': 0,
                'total_failed': len(partner_summaries)
            }
    
    async def get_available_partners(self) -> List[str]:
        """获取可用的Partner列表"""
        try:
            partners = await self.db.get_available_partners()
            # 添加特殊的ALL选项
            if 'ALL' not in partners:
                partners.insert(0, 'ALL')
            return partners
        except Exception as e:
            logger.error(f"❌ 获取Partner列表失败: {e}")
            return ['ALL']
    
    async def get_partner_preview(self, partner_name: str = "ALL",
                                start_date: datetime = None,
                                end_date: datetime = None) -> Dict[str, Any]:
        """获取Partner数据预览"""
        try:
            # 设置默认日期范围
            if not end_date:
                end_date = datetime.now()
            if not start_date:
                start_date = end_date - timedelta(days=7)
            
            # 获取汇总数据
            partner_summaries = await self.db.get_partner_summary(partner_name, start_date, end_date)
            
            # 获取最近的一些转化记录作为预览
            df = await self.db.get_conversion_dataframe(partner_name, start_date, end_date)
            
            preview_data = []
            if not df.empty:
                # 取前10条记录作为预览
                preview_df = df.head(10)
                preview_data = preview_df.to_dict('records')
            
            return {
                'success': True,
                'partner_name': partner_name,
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'total_records': len(df),
                'total_amount': df['Sale Amount (USD)'].sum() if not df.empty else 0,
                'partner_summaries': [summary.to_dict() for summary in partner_summaries],
                'preview_data': preview_data
            }
            
        except Exception as e:
            logger.error(f"❌ 获取Partner预览失败: {e}")
            return {
                'success': False,
                'error': str(e),
                'partner_name': partner_name
            }
    
    async def health_check(self) -> Dict[str, Any]:
        """健康检查"""
        try:
            # 检查数据库连接
            db_health = await self.db.health_check()
            
            # 检查输出目录
            output_dir_exists = self.output_dir.exists()
            output_dir_writable = os.access(self.output_dir, os.W_OK) if output_dir_exists else False
            
            return {
                'status': 'healthy' if db_health['status'] == 'healthy' else 'unhealthy',
                'database': db_health,
                'output_directory': {
                    'exists': output_dir_exists,
                    'writable': output_dir_writable,
                    'path': str(self.output_dir)
                },
                'components': {
                    'feishu_uploader': 'available',
                    'email_sender': 'available'
                },
                'timestamp': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"❌ 健康检查失败: {e}")
            return {
                'status': 'unhealthy',
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
    
    async def cleanup(self):
        """清理资源"""
        try:
            await self.db.close_pool()
            logger.info("✅ 报表生成器资源清理完成")
        except Exception as e:
            logger.error(f"❌ 清理资源失败: {e}") 