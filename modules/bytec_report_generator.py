#!/usr/bin/env python3
"""
ByteC 报表生成器
专门用于生成 ByteC 公司格式的转换报告
与标准 Partner 报表格式不同，包含完整数据和特殊的汇总格式
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from utils.logger import print_step
import config

class ByteCReportGenerator:
    """ByteC 报表生成器类"""
    
    def __init__(self, api_secret=None):
        self.api_secret = api_secret
        self.platform_name = self._get_platform_name(api_secret)
        
    def _get_platform_name(self, api_secret):
        """根据 API Secret 获取平台名称"""
        if api_secret and api_secret in config.API_SECRET_TO_PLATFORM:
            return config.API_SECRET_TO_PLATFORM[api_secret]
        return "Unknown Platform"
    
    def generate_bytec_report(self, raw_data, start_date, end_date, output_dir=None):
        """
        生成 ByteC 格式的报表
        
        Args:
            raw_data: 原始转换数据（DataFrame 或 JSON 数据）
            start_date: 开始日期
            end_date: 结束日期
            output_dir: 输出目录
            
        Returns:
            str: 生成的报表文件路径
        """
        print_step("ByteC报表生成", "开始生成 ByteC 格式的汇总报表...")
        
        if output_dir is None:
            output_dir = config.OUTPUT_DIR
        
        # 步骤1: 数据预处理
        df = self._prepare_data(raw_data)
        
        # 步骤2: 按 offer_name 分组汇总
        summary_data = self._create_offer_summary(df)
        
        # 步骤3: 生成 Excel 文件
        output_path = self._generate_excel_report(summary_data, start_date, end_date, output_dir, raw_data)
        
        print_step("ByteC报表完成", f"成功生成 ByteC 报表: {output_path}")
        return output_path
    
    def _prepare_data(self, raw_data):
        """数据预处理"""
        print_step("数据预处理", "正在处理原始数据...")
        
        # 如果是 JSON 格式，提取 data 部分
        if isinstance(raw_data, dict):
            # 检查是否为多API合并数据
            if 'data' in raw_data and 'conversions' in raw_data['data']:
                # 多API模式的数据结构
                conversion_records = raw_data['data']['conversions']
                print_step("多API数据检测", f"检测到多API合并数据，包含 {len(conversion_records)} 条记录")
                
                # 输出API数据分布信息
                if 'merge_info' in raw_data['data']:
                    merge_info = raw_data['data']['merge_info']
                    print_step("API数据分布", f"数据来源: {merge_info.get('api_breakdown', {})}")
                
            elif 'data' in raw_data and 'data' in raw_data['data']:
                # 单API模式的标准结构
                conversion_records = raw_data['data']['data']
            elif 'data' in raw_data:
                conversion_records = raw_data['data']
            else:
                conversion_records = [raw_data]
            df = pd.DataFrame(conversion_records)
        elif isinstance(raw_data, list):
            df = pd.DataFrame(raw_data)
        elif isinstance(raw_data, pd.DataFrame):
            df = raw_data.copy()
        else:
            raise ValueError("不支持的数据格式")
        
        # 确保必需的字段存在
        required_fields = ['offer_name', 'sale_amount', 'conversion_id', 'aff_sub1']
        missing_fields = [field for field in required_fields if field not in df.columns]
        if missing_fields:
            print_step("数据检查警告", f"缺少必需字段: {missing_fields}")
        
        # 数据类型转换
        if 'sale_amount' in df.columns:
            df['sale_amount'] = pd.to_numeric(df['sale_amount'], errors='coerce').fillna(0)
        
        # 处理 payout 相关字段
        payout_fields = ['payout', 'base_payout', 'bonus_payout']
        for field in payout_fields:
            if field in df.columns:
                df[field] = pd.to_numeric(df[field], errors='coerce').fillna(0)
        
        print_step("数据预处理完成", f"处理了 {len(df)} 条记录，包含 {len(df.columns)} 个字段")
        return df
    
    def _create_offer_summary(self, df):
        """按 offer_name + source 组合创建汇总数据"""
        print_step("数据汇总", "正在按 offer_name + source 组合进行数据汇总...")
        
        if 'offer_name' not in df.columns:
            print_step("汇总警告", "offer_name 字段不存在，无法进行汇总")
            return pd.DataFrame()
        
        if 'aff_sub1' not in df.columns:
            print_step("汇总警告", "aff_sub1 字段不存在，无法进行source分组")
            return pd.DataFrame()
        
        # 按 offer_name + aff_sub1 组合分组汇总
        summary_list = []
        
        for (offer_name, source), group in df.groupby(['offer_name', 'aff_sub1']):
            # 计算汇总数据
            sales_amount = group['sale_amount'].sum() if 'sale_amount' in group.columns else 0
            
            # 计算 Estimated Earning (payout 总和)
            estimated_earning = 0
            if 'payout' in group.columns:
                estimated_earning = group['payout'].sum()
            
            # 转换数量（conversion_id 的数量）
            conversions = len(group) if 'conversion_id' in group.columns else 0
            
            # 获取该组合的 Partner（source已经在分组中确定）
            # 对于ByteC报告，显示实际的原始partner而不是总是显示"ByteC"
            actual_partner = self._get_actual_partner_for_source(source)
            
            # 获取该组合的平台名称（多API模式下从记录中获取）
            platform_name = self._get_platform_for_group(group)
            
            # 计算平均佣金率 (Avg. Commission Rate)
            avg_commission_rate = 0.0
            if sales_amount > 0:
                avg_commission_rate = (estimated_earning / sales_amount) * 100
            
            # 获取广告主佣金率 (Adv Commission Rate)
            adv_commission_rate = config.get_adv_commission_rate(platform_name, avg_commission_rate)
            
            # 计算广告主佣金 (Adv Commission)
            adv_commission = sales_amount * (adv_commission_rate / 100.0)
            
            # 获取发布商佣金率 (Pub Commission Rate)
            pub_commission_rate = config.get_pub_commission_rate(actual_partner, offer_name)
            
            # 计算发布商佣金 (Pub Commission)
            pub_commission = sales_amount * (pub_commission_rate / 100.0)
            
            # 计算ByteC佣金 (ByteC Commission)
            bytec_commission = sales_amount * ((adv_commission_rate - pub_commission_rate) / 100.0)
            
            # 计算ByteC ROI - 新公式: 1 + (Adv Commission - Pub Commission) / Pub Commission
            bytec_roi = 0.0
            if pub_commission > 0:
                bytec_roi = (1 + (adv_commission - pub_commission) / pub_commission) * 100
            else:
                # 如果发布商佣金为0，ROI设为0（避免除零错误）
                bytec_roi = 0.0
            
            summary_list.append({
                'Offer Name': offer_name,
                'Sales Amount': sales_amount,
                'Estimated Earning': estimated_earning, 
                'Partner': actual_partner,
                'Platform': platform_name,
                'Source': source,
                'Conversions': conversions,
                'Avg. Commission Rate': round(avg_commission_rate, 2),
                'Adv Commission Rate': round(adv_commission_rate, 2),
                'Adv Commission': round(adv_commission, 2),
                'Pub Commission Rate': round(pub_commission_rate, 2),
                'Pub Commission': round(pub_commission, 2),
                'ByteC Commission': round(bytec_commission, 2),
                'ByteC ROI': round(bytec_roi, 2)
            })
        
        summary_df = pd.DataFrame(summary_list)
        
        # 按 Offer Name 升序排列
        if len(summary_df) > 0:
            summary_df = summary_df.sort_values('Offer Name', ascending=True)
        
        print_step("汇总完成", f"生成了 {len(summary_df)} 个 Offer + Source 组合的汇总数据")
        return summary_df
    
    def _generate_excel_report(self, summary_data, start_date, end_date, output_dir, raw_data=None):
        """生成 Excel 报表文件"""
        print_step("Excel生成", "正在生成 ByteC Excel 报表...")
        
        # 生成文件名
        filename = config.BYTEC_REPORT_TEMPLATE.format(
            start_date=start_date,
            end_date=end_date
        )
        output_path = os.path.join(output_dir, filename)
        
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        
        # 创建工作簿
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建以日期范围命名的工作表
        if start_date == end_date:
            sheet_name = start_date
        else:
            sheet_name = config.BYTEC_SHEET_NAME_TEMPLATE.format(
                start_date=start_date,
                end_date=end_date
            )
        
        # 确保 Sheet 名称符合 Excel 规范
        sheet_name = self._clean_sheet_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        
        # 写入数据
        if len(summary_data) > 0:
            # 写入标题行和数据
            for r in dataframe_to_rows(summary_data, index=False, header=True):
                cleaned_row = self._clean_row_data(r)
                ws.append(cleaned_row)
            
            # 添加汇总行
            self._add_total_row(ws, summary_data)
            
            # 设置货币格式
            self._apply_currency_formatting(ws, summary_data)
        else:
            # 如果没有数据，写入标题行
            headers = ['Offer Name', 'Sales Amount', 'Estimated Earning', 'Partner', 'Platform', 'Source', 'Conversions', 'Avg. Commission Rate', 'Adv Commission Rate', 'Adv Commission', 'Pub Commission Rate', 'Pub Commission', 'ByteC Commission', 'ByteC ROI']
            ws.append(headers)
        
        # 保存文件
        wb.save(output_path)
        
        # 输出统计信息
        total_sales = summary_data['Sales Amount'].sum() if len(summary_data) > 0 else 0
        total_earning = summary_data['Estimated Earning'].sum() if len(summary_data) > 0 else 0
        total_conversions = summary_data['Conversions'].sum() if len(summary_data) > 0 else 0
        
        print_step("Excel统计", f"总销售额: ${total_sales:,.2f}, 总预计收入: ${total_earning:,.2f}, 总转换数: {total_conversions}")
        
        # 生成详细的报告摘要
        summary_report = self._generate_report_summary(summary_data, start_date, end_date, raw_data)
        print_step("报告摘要", summary_report)
        
        return output_path
    
    def _get_actual_partner_for_source(self, source):
        """获取source对应的实际partner，忽略ByteC的配置"""
        import re
        
        # 遍历所有partner配置，但跳过ByteC
        for partner, partner_config in config.PARTNER_SOURCES_MAPPING.items():
            if partner == "ByteC":  # 跳过ByteC配置
                continue
                
            # 先检查sources列表
            if source in partner_config.get('sources', []):
                return partner
            
            # 再检查正则表达式模式
            pattern = partner_config.get('pattern', '')
            if pattern and re.match(pattern, source):
                return partner
        
        # 如果没有匹配到具体的partner，返回source本身作为partner
        return source
    
    def _get_platform_for_group(self, group):
        """获取该组数据的平台名称"""
        # 在多API模式下，尝试从记录中获取api_source字段
        if 'api_source' in group.columns and not group['api_source'].empty:
            # 获取该组的api_source（应该都是相同的）
            api_source = group['api_source'].iloc[0]
            return api_source
        
        # 单API模式或没有api_source字段时，使用初始化时的platform_name
        return self.platform_name
    
    def _add_total_row(self, worksheet, summary_data):
        """添加汇总行到工作表底部"""
        if len(summary_data) == 0:
            return
        
        # 计算汇总数据
        total_sales = summary_data['Sales Amount'].sum()
        total_earning = summary_data['Estimated Earning'].sum()
        total_conversions = summary_data['Conversions'].sum()
        
        # 计算总体平均佣金率
        total_avg_commission_rate = 0.0
        if total_sales > 0:
            total_avg_commission_rate = (total_earning / total_sales) * 100
        
        # 计算新佣金栏位的汇总
        total_adv_commission = summary_data['Adv Commission'].sum() if 'Adv Commission' in summary_data.columns else 0
        total_pub_commission = summary_data['Pub Commission'].sum() if 'Pub Commission' in summary_data.columns else 0
        total_bytec_commission = summary_data['ByteC Commission'].sum() if 'ByteC Commission' in summary_data.columns else 0
        
        # 计算总体ByteC ROI - 新公式: 1 + (Total Adv Commission - Total Pub Commission) / Total Pub Commission
        total_bytec_roi = 0.0
        if total_pub_commission > 0:
            total_bytec_roi = (1 + (total_adv_commission - total_pub_commission) / total_pub_commission) * 100
        else:
            total_bytec_roi = 0.0
        
        # 添加空行作为分隔
        worksheet.append([])
        
        # 添加汇总行
        total_row = [
            'TOTAL',  # Offer Name
            total_sales,  # Sales Amount
            total_earning,  # Estimated Earning
            '',  # Partner (留空)
            '',  # Platform (留空)
            '',  # Source (留空)
            total_conversions,  # Conversions
            round(total_avg_commission_rate, 2),  # Avg. Commission Rate
            '',  # Adv Commission Rate (留空 - 因为是不同平台的混合)
            round(total_adv_commission, 2),  # Adv Commission
            '',  # Pub Commission Rate (留空 - 因为是不同Partner的混合)
            round(total_pub_commission, 2),  # Pub Commission
            round(total_bytec_commission, 2),  # ByteC Commission
            round(total_bytec_roi, 2)  # ByteC ROI
        ]
        
        cleaned_total_row = self._clean_row_data(total_row)
        worksheet.append(cleaned_total_row)
        
        print_step("汇总行", f"已添加汇总行 - 总销售: ${total_sales:,.2f}, 总收入: ${total_earning:,.2f}, 总转换: {total_conversions}")
        print_step("佣金汇总", f"广告主佣金: ${total_adv_commission:,.2f}, 发布商佣金: ${total_pub_commission:,.2f}, ByteC佣金: ${total_bytec_commission:,.2f}, ByteC ROI: {total_bytec_roi:.2f}%")
    
    def _generate_report_summary(self, summary_data, start_date, end_date, raw_data=None):
        """生成详细的报告摘要"""
        if len(summary_data) == 0:
            return "报告摘要: 无数据"
        
        # 基本统计信息
        total_offers = len(summary_data)
        total_sales = summary_data['Sales Amount'].sum()
        total_earning = summary_data['Estimated Earning'].sum()
        total_conversions = summary_data['Conversions'].sum()
        
        # 获取前5个Offer的信息
        top_offers = summary_data.head(5)
        
        # 构建摘要文本
        summary_lines = [
            f"📊 ByteC报告摘要 ({start_date} 至 {end_date})",
            f"=" * 50,
        ]
        
        # 添加API来源信息（多API模式）
        if raw_data and isinstance(raw_data, dict) and 'data' in raw_data:
            data_section = raw_data['data']
            if 'merge_info' in data_section:
                merge_info = data_section['merge_info']
                summary_lines.extend([
                    f"🔗 数据来源 (多API模式):",
                    f"  • 总API数量: {merge_info.get('total_apis', 0)}",
                    f"  • 成功API数量: {merge_info.get('successful_apis', 0)}",
                    f"  • API数据分布: {merge_info.get('api_breakdown', {})}",
                    ""
                ])
            elif 'api_sources' in data_section:
                api_sources = data_section['api_sources']
                summary_lines.extend([
                    f"🔗 数据来源API: {', '.join(api_sources)}",
                    ""
                ])
        
        summary_lines.extend([
            f"📈 总体数据:",
            f"  • Offer数量: {total_offers}",
            f"  • 总销售额: ${total_sales:,.2f}",
            f"  • 总预计收入: ${total_earning:,.2f}",
            f"  • 总转换数: {total_conversions:,}",
            f"  • 平均每个Offer销售额: ${total_sales/total_offers:,.2f}",
            "",
            f"🏆 Top {min(5, total_offers)} Offers (按销售额排序):"
        ])
        
        for i, (_, row) in enumerate(top_offers.iterrows(), 1):
            summary_lines.append(
                f"  {i}. {row['Offer Name']}: ${row['Sales Amount']:,.2f} "
                f"(收入: ${row['Estimated Earning']:,.2f}, 转换: {row['Conversions']})"
            )
        
        # 添加平台和伙伴统计
        if 'Platform' in summary_data.columns:
            platforms = summary_data['Platform'].value_counts()
            if len(platforms) > 0:
                summary_lines.extend([
                    "",
                    f"🔗 平台分布:",
                    f"  • {', '.join([f'{platform}: {count}个offer' for platform, count in platforms.items()])}"
                ])
        
        # 添加Partner分布统计
        if 'Partner' in summary_data.columns:
            partners = summary_data['Partner'].value_counts()
            if len(partners) > 0:
                summary_lines.extend([
                    "",
                    f"👥 Partner分布:",
                    f"  • {', '.join([f'{partner}: {count}个offer' for partner, count in partners.items()])}"
                ])
        
        # 添加转换率信息
        if total_conversions > 0:
            avg_earning_per_conversion = total_earning / total_conversions
            summary_lines.extend([
                "",
                f"💰 效率指标:",
                f"  • 每转换平均收入: ${avg_earning_per_conversion:.2f}",
                f"  • 收入与销售比率: {(total_earning/total_sales*100) if total_sales > 0 else 0:.2f}%"
            ])
        
        return "\n".join(summary_lines)
    
    def _apply_currency_formatting(self, worksheet, data):
        """为货币字段和百分比字段应用格式"""
        if len(data) == 0:
            return
        
        # 查找各列的索引
        columns = list(data.columns)
        
        # 货币格式栏位
        currency_columns = []
        # 百分比格式栏位  
        percentage_columns = []
        
        for i, column_name in enumerate(columns):
            excel_col_index = i + 1  # Excel 列索引从1开始
            
            if column_name in ['Sales Amount', 'Estimated Earning', 'Adv Commission', 'Pub Commission', 'ByteC Commission']:
                currency_columns.append(excel_col_index)
            elif column_name in ['Avg. Commission Rate', 'Adv Commission Rate', 'Pub Commission Rate', 'ByteC ROI']:
                percentage_columns.append(excel_col_index)
        
        # 计算总行数（包括数据行、空行和汇总行）
        total_rows = len(data) + 3
        
        # 应用货币格式 ($0.00)
        for col_index in currency_columns:
            for row in range(2, total_rows + 1):  # 从第2行开始，跳过标题行
                cell = worksheet.cell(row=row, column=col_index)
                if cell.value is not None and str(cell.value).replace('.', '').replace('-', '').isdigit():
                    try:
                        cell.number_format = '"$"#,##0.00'
                    except:
                        pass
        
        # 应用百分比格式 (0.00%)
        for col_index in percentage_columns:
            for row in range(2, total_rows + 1):  # 从第2行开始，跳过标题行
                cell = worksheet.cell(row=row, column=col_index)
                if cell.value is not None and str(cell.value).replace('.', '').replace('-', '').isdigit():
                    try:
                        cell.number_format = '0.00%'
                        # 将百分比值转换为小数形式 (Excel要求)
                        # 所有百分比值都需要除以100转换为小数
                        if isinstance(cell.value, (int, float)):
                            cell.value = cell.value / 100.0
                    except:
                        pass
        
        print_step("格式设置", f"已为 {len(currency_columns)} 个货币栏位设置美元格式，{len(percentage_columns)} 个栏位设置百分比格式")
        
        # 添加条件格式 - ByteC ROI为负数时标红
        if 'ByteC ROI' in columns:
            roi_col_index = columns.index('ByteC ROI') + 1  # Excel列索引从1开始
            roi_col_letter = chr(ord('A') + roi_col_index - 1)  # 转换为Excel列字母
            
            # 为ByteC ROI列添加条件格式（负数标红）
            from openpyxl.styles import PatternFill, Font
            from openpyxl.formatting.rule import CellIsRule
            
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            red_font = Font(color='CC0000')
            
            # 应用条件格式到数据范围（不包括标题行）
            data_range = f'{roi_col_letter}2:{roi_col_letter}{len(data) + 1}'
            rule = CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font)
            worksheet.conditional_formatting.add(data_range, rule)
            
            print(f"      ✓ ByteC ROI负数标红格式已应用到范围: {data_range}")
    
    def _clean_sheet_name(self, name):
        """清理 Sheet 名称，确保符合 Excel 规范"""
        import re
        
        # 移除或替换不允许的字符
        clean_name = re.sub(r'[\\/*?[\]:]+', '_', str(name))
        
        # 限制长度
        if len(clean_name) > 31:
            clean_name = clean_name[:28] + "..."
        
        # 确保不为空
        if not clean_name:
            clean_name = "ByteC_Report"
        
        return clean_name
    
    def _clean_row_data(self, row):
        """清理行数据，确保 Excel 兼容性"""
        import re
        
        cleaned_row = []
        for cell in row:
            if cell is None:
                cleaned_row.append(None)
            elif isinstance(cell, (int, float)):
                cleaned_row.append(cell)
            else:
                # 清理字符串中的特殊字符
                cell_str = str(cell)
                cleaned_str = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', cell_str)
                cleaned_str = cleaned_str.strip()
                cleaned_row.append(cleaned_str)
        
        return cleaned_row 