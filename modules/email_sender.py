#!/usr/bin/env python3
"""
邮件发送模块
负责发送转换报告邮件
"""

import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from utils.logger import print_step
import config
import pandas as pd

class EmailSender:
    """邮件发送器"""
    
    def __init__(self):
        self.sender = config.EMAIL_SENDER
        self.default_receivers = config.EMAIL_RECEIVERS
        self.partner_email_mapping = getattr(config, 'PARTNER_EMAIL_MAPPING', config.PUB_EMAIL_MAPPING)  # 新的Partner配置
        self.partner_email_enabled = getattr(config, 'PARTNER_EMAIL_ENABLED', config.PUB_EMAIL_ENABLED)  # 新的Partner配置
        # 保持向后兼容性
        self.pub_email_mapping = self.partner_email_mapping  # 兼容性别名
        self.pub_email_enabled = self.partner_email_enabled  # 兼容性别名
        self.auto_cc_email = getattr(config, 'EMAIL_AUTO_CC', None)  # 自动抄送邮箱
        self.password = config.EMAIL_PASSWORD
        self.smtp_server = config.SMTP_SERVER
        self.smtp_port = config.SMTP_PORT
        self.enable_tls = config.EMAIL_ENABLE_TLS
        self.include_attachments = config.EMAIL_INCLUDE_ATTACHMENTS
        self.include_feishu_links = config.EMAIL_INCLUDE_FEISHU_LINKS
        self.subject_template = config.EMAIL_SUBJECT_TEMPLATE
    
    def send_partner_reports(self, partner_summary, feishu_upload_result=None, report_date=None, start_date=None):
        """
        按Partner分别发送转换报告邮件
        
        Args:
            partner_summary: Partner汇总数据字典，格式：
                {
                    'partner_name': {
                        'records': 数量,
                        'amount_formatted': '$金额',
                        'file_path': '文件路径',
                        'sources': ['source1', 'source2', ...],
                        'sources_count': 数量
                    }
                }
            feishu_upload_result: 飞书上传结果
            report_date: 报告日期（结束日期，用于邮件标题和内容）
            start_date: 开始日期（用于邮件中的日期范围显示）
        
        Returns:
            dict: 发送结果汇总
        """
        print_step("Partner邮件发送", "开始按Partner分别发送转换报告邮件")
        
        # 检查配置
        if self.password == "your_gmail_app_password_here":
            error_msg = "邮件密码未配置，请在config.py中设置EMAIL_PASSWORD"
            print_step("配置错误", f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
        
        results = {
            'success': True,
            'total_sent': 0,
            'total_failed': 0,
            'partner_results': {}
        }
        
        for partner_name, partner_data in partner_summary.items():
            try:
                # 检查是否在目标Partner列表中
                if not config.is_partner_enabled(partner_name):
                    print_step(f"Partner邮件-{partner_name}", f"⚠️ 跳过Partner '{partner_name}' (不在处理范围内)")
                    results['partner_results'][partner_name] = {'success': True, 'skipped': True, 'reason': '不在处理范围内'}
                    continue
                
                # 检查该Partner是否启用邮件发送
                if not self.partner_email_enabled.get(partner_name, False):
                    print_step(f"Partner邮件-{partner_name}", f"⚠️ 邮件发送已关闭，跳过")
                    results['partner_results'][partner_name] = {'success': True, 'skipped': True, 'reason': '邮件发送已关闭'}
                    continue
                
                # 获取该Partner的收件人
                receivers = self.partner_email_mapping.get(partner_name, self.default_receivers)
                
                # 准备该Partner的邮件数据
                email_data = self._prepare_partner_email_data(partner_name, partner_data, report_date, start_date)
                
                # 获取该Partner的飞书文件信息
                partner_feishu_info = self._get_partner_feishu_info(partner_name, feishu_upload_result)
                
                # 发送邮件
                result = self._send_single_partner_email(
                    partner_name, 
                    email_data, 
                    [partner_data['file_path']], 
                    receivers,
                    partner_feishu_info,
                    report_date
                )
                
                results['partner_results'][partner_name] = result
                if result['success'] and not result.get('skipped'):
                    results['total_sent'] += 1
                elif result.get('skipped'):
                    # 跳过的不计入失败，但也不计入成功发送
                    pass
                else:
                    results['total_failed'] += 1
                    results['success'] = False
                    
            except Exception as e:
                error_msg = f"Partner {partner_name} 邮件发送失败: {str(e)}"
                print_step("Partner邮件失败", f"❌ {error_msg}")
                results['partner_results'][partner_name] = {'success': False, 'error': error_msg}
                results['total_failed'] += 1
                results['success'] = False
        
        if results['success']:
            print_step("Partner邮件完成", f"✅ 成功发送 {results['total_sent']} 个Partner报告邮件")
        else:
            print_step("Partner邮件部分失败", f"⚠️ 发送完成：成功 {results['total_sent']} 个，失败 {results['total_failed']} 个")
        
        return results
    
    def send_report_email(self, report_data, file_paths=None, feishu_upload_result=None):
        """
        发送转换报告邮件（兼容性保留，建议使用send_pub_reports）
        
        Args:
            report_data: 报告数据字典
            file_paths: Excel文件路径列表
            feishu_upload_result: 飞书上传结果
        
        Returns:
            dict: 发送结果
        """
        print_step("邮件发送", "开始发送转换报告邮件")
        
        # 检查配置
        if self.password == "your_gmail_app_password_here":
            error_msg = "邮件密码未配置，请在config.py中设置EMAIL_PASSWORD"
            print_step("配置错误", f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
        
        try:
            # 创建邮件对象
            msg = self._create_email_message(report_data, file_paths, feishu_upload_result, self.default_receivers)
            
            # 发送邮件
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                if self.enable_tls:
                    server.starttls()
                
                server.login(self.sender, self.password)
                text = msg.as_string()
                server.sendmail(self.sender, self.default_receivers, text)
            
            print_step("邮件发送成功", f"✅ 邮件已发送给 {', '.join(self.default_receivers)}")
            return {
                'success': True,
                'recipients': self.default_receivers,
                'attachments_count': len(file_paths) if file_paths and self.include_attachments else 0
            }
            
        except Exception as e:
            error_msg = f"邮件发送失败: {str(e)}"
            print_step("邮件发送失败", f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
    
    def _send_single_partner_email(self, partner_name, email_data, file_paths, receivers, feishu_info, report_date=None):
        """发送单个Partner的邮件"""
        try:
            # 添加抄送邮箱（从配置读取）
            cc_email = self.auto_cc_email
            all_recipients = receivers.copy()
            if cc_email and cc_email not in receivers:
                all_recipients.append(cc_email)
            
            # 创建邮件对象
            msg = self._create_partner_email_message(partner_name, email_data, file_paths, receivers, feishu_info, report_date, cc_email)
            
            # 发送邮件
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                if self.enable_tls:
                    server.starttls()
                
                server.login(self.sender, self.password)
                text = msg.as_string()
                server.sendmail(self.sender, all_recipients, text)
            
            cc_info = f" (抄送: {cc_email})" if cc_email else ""
            print_step(f"Partner邮件-{partner_name}", f"✅ 邮件已发送给 {', '.join(receivers)}{cc_info}")
            return {
                'success': True,
                'recipients': receivers,
                'cc_recipients': [cc_email] if cc_email else [],
                'attachments_count': len(file_paths) if file_paths and self.include_attachments else 0
            }
            
        except Exception as e:
            error_msg = f"发送失败: {str(e)}"
            print_step(f"Partner邮件-{partner_name}", f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
    
    def _prepare_partner_email_data(self, partner_name, partner_data, report_date=None, start_date=None):
        """准备Partner邮件数据"""
        if report_date is None:
            report_date = datetime.now().strftime("%Y-%m-%d")
        if start_date is None:
            start_date = report_date
        
        file_path = partner_data.get('file_path')
        
        # 从Excel文件中计算真实的销售总额
        real_total_amount = self._calculate_sales_amount_from_excel(file_path)
        
        # 计算Sources统计信息
        sources_statistics = self._calculate_sources_statistics_from_excel(file_path)
        
        return {
            'partner_name': partner_name,
            'total_records': partner_data.get('records', 0),
            'total_amount': real_total_amount,
            'start_date': start_date,
            'end_date': report_date,
            'report_date': report_date,
            'main_file': os.path.basename(file_path) if file_path else '',
            'file_path': file_path,  # 添加完整文件路径，供ByteC邮件模板使用
            'sources': partner_data.get('sources', []),
            'sources_count': partner_data.get('sources_count', 0),
            'sources_statistics': sources_statistics
        }
    
    def _calculate_sales_amount_from_excel(self, file_path):
        """从Excel文件中计算Sales Amount总额（包含所有sheets）"""
        try:
            if not file_path or not os.path.exists(file_path):
                print_step("金额计算", f"⚠️ 文件不存在: {file_path}")
                return '$0.00'
            
            import openpyxl
            
            # 使用openpyxl读取所有sheets
            wb = openpyxl.load_workbook(file_path, read_only=True)
            total_amount = 0.0
            sheet_details = []
            
            print_step("金额计算", f"📊 正在计算 {os.path.basename(file_path)} 的销售总额（包含所有sheets）...")
            
            for sheet_name in wb.sheetnames:
                try:
                    # 读取该sheet的数据
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # 支持多种可能的销售金额列名
                    sales_amount_col = None
                    possible_col_names = ['Sales Amount', 'sale_amount', 'Sale Amount', 'sales_amount', 'SALES_AMOUNT']
                    
                    for col_name in possible_col_names:
                        if col_name in df.columns:
                            sales_amount_col = col_name
                            break
                    
                    if sales_amount_col and len(df) > 0:
                        sheet_total = df[sales_amount_col].sum()
                        total_amount += sheet_total
                        sheet_details.append(f"  - {sheet_name}: ${sheet_total:.2f}")
                        print_step("金额计算", f"📋 Sheet '{sheet_name}': ${sheet_total:.2f} ({len(df)} 条记录，使用列'{sales_amount_col}')")
                    else:
                        sheet_details.append(f"  - {sheet_name}: $0.00 (无数据或无销售金额列)")
                        print_step("金额计算", f"⚠️ Sheet '{sheet_name}': 无销售金额列或无数据")
                
                except Exception as e:
                    print_step("金额计算", f"⚠️ 处理Sheet '{sheet_name}' 失败: {str(e)}")
                    sheet_details.append(f"  - {sheet_name}: 计算失败")
            
            wb.close()
            
            formatted_amount = f"${total_amount:.2f}"
            print_step("金额计算", f"💰 {os.path.basename(file_path)} 总销售额: {formatted_amount}")
            print_step("金额详情", f"各Sheet明细:\n" + "\n".join(sheet_details))
            
            return formatted_amount
                
        except Exception as e:
            print_step("金额计算", f"❌ 计算金额失败 {os.path.basename(file_path)}: {str(e)}")
            return '$0.00'
    
    def _calculate_sources_statistics_from_excel(self, file_path):
        """从Excel文件中计算各Sources的统计信息"""
        try:
            if not file_path or not os.path.exists(file_path):
                return []
            
            import openpyxl
            
            # 使用openpyxl读取多个sheets
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sources_stats = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 跳过第一行标题，计算记录数
                row_count = ws.max_row - 1 if ws.max_row > 1 else 0
                
                # 读取该sheet的数据来计算销售金额
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # 支持多种可能的销售金额列名
                sales_amount_col = None
                possible_col_names = ['Sales Amount', 'sale_amount', 'Sale Amount', 'sales_amount', 'SALES_AMOUNT']
                
                for col_name in possible_col_names:
                    if col_name in df.columns:
                        sales_amount_col = col_name
                        break
                
                if sales_amount_col and len(df) > 0:
                    sales_amount = df[sales_amount_col].sum()
                    formatted_amount = f"${sales_amount:.2f}"
                else:
                    formatted_amount = '$0.00'
                
                sources_stats.append({
                    'source_name': sheet_name,
                    'records': row_count,
                    'sales_amount': formatted_amount
                })
            
            wb.close()
            return sources_stats
            
        except Exception as e:
            print_step("Sources统计", f"❌ 计算Sources统计失败 {os.path.basename(file_path)}: {str(e)}")
            return []
    
    def _get_partner_feishu_info(self, partner_name, feishu_upload_result):
        """获取该Partner的飞书文件信息"""
        if not feishu_upload_result or not feishu_upload_result.get('uploaded_files'):
            return None
        
        # 查找该Partner对应的飞书文件
        for file_info in feishu_upload_result['uploaded_files']:
            filename = file_info.get('filename', '')
            if filename.startswith(partner_name):
                return {
                    'success': True,
                    'uploaded_files': [file_info]
                }
        
        return None
    
    def _create_partner_email_message(self, partner_name, email_data, file_paths, receivers, feishu_info, report_date=None, cc_email=None):
        """创建Partner邮件消息"""
        # 生成邮件主题 - 使用附件文档名称（不含.xlsx扩展名）
        main_file = email_data.get('main_file', f'{partner_name}_ConversionReport_{report_date or datetime.now().strftime("%Y-%m-%d")}.xlsx')
        # 移除.xlsx扩展名
        subject = main_file.replace('.xlsx', '') if main_file.endswith('.xlsx') else main_file
        
        # 创建邮件对象
        msg = MIMEMultipart()
        msg['From'] = self.sender
        msg['To'] = ", ".join(receivers)
        if cc_email:
            msg['Cc'] = cc_email
        msg['Subject'] = subject
        
        # 生成邮件正文
        body = self._generate_partner_email_body(partner_name, email_data, feishu_info)
        msg.attach(MIMEText(body, 'html'))
        
        # 添加附件
        if file_paths and self.include_attachments:
            for file_path in file_paths:
                if os.path.exists(file_path):
                    self._attach_file(msg, file_path)
        
        return msg
    
    def _create_email_message(self, report_data, file_paths=None, feishu_upload_result=None, receivers=None):
        """创建邮件消息（兼容性保留）"""
        if receivers is None:
            receivers = self.default_receivers
            
        # 生成邮件主题
        today = datetime.now().strftime("%Y-%m-%d")
        subject = self.subject_template.format(date=today)
        
        # 创建邮件对象
        msg = MIMEMultipart()
        msg['From'] = self.sender
        msg['To'] = ", ".join(receivers)
        msg['Subject'] = subject
        
        # 生成邮件正文
        body = self._generate_email_body(report_data, feishu_upload_result)
        msg.attach(MIMEText(body, 'html'))
        
        # 添加附件
        if file_paths and self.include_attachments:
            for file_path in file_paths:
                if os.path.exists(file_path):
                    self._attach_file(msg, file_path)
        
        return msg
    
    def _load_html_template(self, template_name):
        """加载HTML模板文件"""
        try:
            template_path = os.path.join('templates', template_name)
            with open(template_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            print_step("模板加载", f"❌ 加载模板失败 {template_name}: {str(e)}")
            return None
    
    def _generate_partner_email_body(self, partner_name, email_data, feishu_info):
        """生成Partner专用邮件正文"""
        # 检查是否为ByteC Partner，使用专用模板
        if self._is_bytec_partner(partner_name):
            return self._generate_bytec_email_body(partner_name, email_data, feishu_info)
        
        report_date = email_data.get('report_date', datetime.now().strftime("%Y-%m-%d"))
        completion_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 从email_data获取Partner专用数据
        total_records = email_data.get('total_records', 0)
        total_amount = email_data.get('total_amount', '$0.00')
        start_date = email_data.get('start_date', report_date)
        end_date = email_data.get('end_date', report_date)
        main_file = email_data.get('main_file', f'{partner_name}_ConversionReport_{report_date}.xlsx')
        sources_statistics = email_data.get('sources_statistics', [])
        
        # 加载HTML模板
        template = self._load_html_template('email_template.html')
        if not template:
            # 如果模板加载失败，使用备用的简单HTML
            return self._generate_fallback_email_body(partner_name, email_data, feishu_info)
        
        # 准备飞书链接部分
        feishu_section = ""
        if feishu_info and self.include_feishu_links:
            if feishu_info.get('success') and feishu_info.get('uploaded_files'):
                feishu_template = self._load_html_template('feishu_section.html')
                if feishu_template:
                    feishu_links = ""
                    for file_info in feishu_info['uploaded_files']:
                        filename = file_info.get('filename', '')
                        
                        # 优先使用飞书返回的URL
                        feishu_url = file_info.get('url')
                        if not feishu_url:
                            # 如果没有url字段，尝试用file_token构造
                            file_token = file_info.get('file_token') or file_info.get('file_id')
                            if file_token:
                                feishu_url = f"https://bytedance.feishu.cn/sheets/{file_token}"
                        
                        if feishu_url:
                            feishu_links += f'<li><a href="{feishu_url}" target="_blank" style="color: #007bff; text-decoration: none; font-weight: bold;">{filename}</a></li>'
                        else:
                            feishu_links += f'<li>{filename} (已上传)</li>'
                    
                    feishu_section = feishu_template.replace('{{feishu_links}}', feishu_links)
        
        # 生成Sources统计HTML和Sources列表
        sources_statistics_html = self._generate_sources_statistics_html(sources_statistics)
        sources_list = self._generate_sources_list(sources_statistics)
        
        # 替换模板中的占位符
        body = template.replace('{{date}}', report_date)
        body = body.replace('{{partner_name}}', partner_name)  # 注意：这里改为partner_name
        body = body.replace('{{start_date}}', start_date)
        body = body.replace('{{end_date}}', end_date)
        body = body.replace('{{total_records}}', str(total_records))
        body = body.replace('{{total_amount}}', total_amount)
        body = body.replace('{{main_file}}', main_file)
        body = body.replace('{{sources_list}}', sources_list)
        body = body.replace('{{sources_statistics}}', sources_statistics_html)
        body = body.replace('{{feishu_section}}', feishu_section)
        body = body.replace('{{completion_time}}', completion_time)
        
        return body
    
    def _generate_sources_statistics_html(self, sources_statistics):
        """生成Sources统计的HTML（新格式：列表形式）"""
        if not sources_statistics:
            return "<p>暂无Sources统计数据</p>"
        
        html_parts = ["<ul style='list-style: none; padding: 0; margin: 0;'>"]
        
        for stat in sources_statistics:
            source_name = stat.get('source_name', 'Unknown')
            records = stat.get('records', 0)
            sales_amount = stat.get('sales_amount', '$0.00')
            
            html_parts.append(f"<li style='margin: 8px 0; padding: 8px; background-color: #ffffff; border: 1px solid #e9ecef; border-radius: 4px;'>")
            html_parts.append(f"<strong>- {source_name}:</strong> ")
            html_parts.append(f"Total Conversion: <strong>{records}</strong> 条, ")
            html_parts.append(f"Total Sales Amount: <span style='color: #28a745; font-weight: bold;'>{sales_amount}</span>")
            html_parts.append("</li>")
        
        html_parts.append("</ul>")
        
        return "".join(html_parts)
    
    def _generate_sources_list(self, sources_statistics):
        """生成Sources列表字符串"""
        if not sources_statistics:
            return "无"
        
        source_names = [stat.get('source_name', 'Unknown') for stat in sources_statistics]
        return ", ".join(source_names)
    
    def _generate_fallback_email_body(self, partner_name, email_data, feishu_info):
        """生成备用的简单邮件正文（当模板加载失败时使用）"""
        report_date = email_data.get('report_date', datetime.now().strftime("%Y-%m-%d"))
        completion_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        total_records = email_data.get('total_records', 0)
        total_amount = email_data.get('total_amount', '$0.00')
        start_date = email_data.get('start_date', report_date)
        end_date = email_data.get('end_date', report_date)
        main_file = email_data.get('main_file', f'{partner_name}_ConversionReport_{report_date}.xlsx')
        sources_statistics = email_data.get('sources_statistics', [])
        
        # 生成Sources列表
        sources_list = self._generate_sources_list(sources_statistics)
        
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <p>Hi {partner_name} Teams,</p>
            <p>{main_file} 如附件，请查收。</p>
            
            <h3 style="color: #1f4e79;">📊 {partner_name} 报告摘要:</h3>
            <ul style="background-color: #f8f9fa; padding: 15px; border-left: 4px solid #007bff;">
                <li><strong>Partner:</strong> {partner_name}</li>
                <li><strong>Date Range:</strong> {start_date} 至 {end_date}</li>
                <li><strong>Total Conversion:</strong> {total_records} 条</li>
                <li><strong>Total Sales Amount:</strong> {total_amount}</li>
                <li><strong>Sources:</strong> {sources_list}</li>
            </ul>
            <div style="background-color: #f8f9fa; padding: 15px; border-left: 4px solid #007bff; margin-top: 15px;">
                {self._generate_sources_statistics_html(sources_statistics)}
            </div>
            
            <h3 style="color: #1f4e79;">📁 附件文件:</h3>
            <ul><li><strong>{main_file}</strong></li></ul>
        """
        
        # 飞书文件链接
        if feishu_info and self.include_feishu_links:
            if feishu_info.get('success') and feishu_info.get('uploaded_files'):
                body += f'<h3 style="color: #1f4e79;">☁️ 飞书文件链接:</h3><ul>'
                for file_info in feishu_info['uploaded_files']:
                    filename = file_info.get('filename', '')
                    
                    # 优先使用飞书返回的URL
                    feishu_url = file_info.get('url')
                    if not feishu_url:
                        # 如果没有url字段，尝试用file_token构造
                        file_token = file_info.get('file_token') or file_info.get('file_id')
                        if file_token:
                            feishu_url = f"https://bytedance.feishu.cn/sheets/{file_token}"
                    
                    if feishu_url:
                        body += f'<li><a href="{feishu_url}" target="_blank" style="color: #007bff; text-decoration: none; font-weight: bold;">{filename}</a></li>'
                    else:
                        body += f'<li>{filename} (已上传)</li>'
                body += "</ul>"
        
        body += f"""
            <p style="margin-top: 30px;"><strong>生成时间:</strong> {completion_time}</p>
            <p style="margin-top: 30px; color: #666;">
                Best regards,<br><strong>AutoReporter Agent</strong>
            </p>
        </body>
        </html>
        """
        
        return body
    
    def _generate_email_body(self, report_data, feishu_upload_result=None):
        """生成邮件正文（兼容性保留）"""
        today = datetime.now().strftime("%Y-%m-%d")
        
        # 基本信息
        total_records = report_data.get('total_records', 0)
        total_amount = report_data.get('total_amount', '$0.00')
        start_date = report_data.get('start_date', today)
        end_date = report_data.get('end_date', today)
        completion_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 构建邮件正文
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <p>Hi Partners,</p>
            
            <p>Conversion Report + {start_date} 如附件，请查收。</p>
            
            <h3 style="color: #1f4e79;">📊 报告摘要:</h3>
            <ul style="background-color: #f8f9fa; padding: 15px; border-left: 4px solid #007bff;">
                <li><strong>日期范围:</strong> {start_date} 至 {end_date}</li>
                <li><strong>总转换数:</strong> {total_records} 条</li>
                <li><strong>Total Sales Amount:</strong> {total_amount}</li>
            </ul>
            
            <h3 style="color: #1f4e79;">📁 生成的文件:</h3>
            <ul>
        """
        
        # 添加文件信息
        pub_files = report_data.get('pub_files', [])
        main_file = report_data.get('main_file', f'Pub_ConversionReport_{today}.xlsx')
        
        body += f"<li><strong>主报告:</strong> {os.path.basename(main_file)}</li>"
        
        for pub_file_info in pub_files:
            if isinstance(pub_file_info, dict):
                filename = pub_file_info.get('filename', '')
                records = pub_file_info.get('records', 0)
                amount = pub_file_info.get('amount', '$0.00')
                body += f"<li><strong>{filename}:</strong> ({records}条, {amount})</li>"
            else:
                # 如果是文件路径字符串
                filename = os.path.basename(pub_file_info)
                body += f"<li><strong>{filename}</strong></li>"
        
        body += "</ul>"
        
        # 飞书上传状态
        if feishu_upload_result:
            if feishu_upload_result.get('success'):
                feishu_status = f"✅ 成功 ({feishu_upload_result.get('success_count', 0)} 个文件)"
                if self.include_feishu_links and feishu_upload_result.get('uploaded_files'):
                    body += f"<h3 style='color: #1f4e79;'>☁️ 飞书文件链接:</h3><ul>"
                    for file_info in feishu_upload_result['uploaded_files']:
                        if file_info.get('url'):
                            body += f"<li><a href='{file_info['url']}'>{file_info['filename']}</a></li>"
                    body += "</ul>"
            else:
                feishu_status = f"❌ 失败 ({feishu_upload_result.get('failed_count', 0)} 个文件)"
        else:
            feishu_status = "未执行"
        
        body += f"""
            <h3 style="color: #1f4e79;">☁️ 飞书上传状态:</h3>
            <p style="background-color: #f8f9fa; padding: 10px; border-radius: 5px;">{feishu_status}</p>
            
            <p style="margin-top: 30px;"><strong>生成时间:</strong> {completion_time}</p>
            
            <p style="margin-top: 30px; color: #666;">
                Best regards,<br>
                <strong>AutoReporter Agent</strong>
            </p>
        </body>
        </html>
        """
        
        return body
    
    def _attach_file(self, msg, file_path):
        """添加附件到邮件"""
        try:
            with open(file_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            filename = os.path.basename(file_path)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {filename}'
            )
            msg.attach(part)
            print_step("附件添加", f"已添加附件: {filename}")
            
        except Exception as e:
            print_step("附件错误", f"⚠️ 无法添加附件 {file_path}: {str(e)}")
    
    def test_connection(self):
        """测试邮件服务器连接"""
        print_step("邮件测试", "正在测试邮件服务器连接...")
        
        if self.password == "your_gmail_app_password_here":
            print_step("配置错误", "❌ 邮件密码未配置")
            return False
        
        try:
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                if self.enable_tls:
                    server.starttls()
                server.login(self.sender, self.password)
            
            print_step("邮件测试", "✅ 邮件服务器连接成功")
            return True
            
        except Exception as e:
            print_step("邮件测试", f"❌ 邮件服务器连接失败: {str(e)}")
            return False

    # 保持向后兼容性的方法别名
    def send_pub_reports(self, pub_summary, feishu_upload_result=None, report_date=None):
        """向后兼容性方法，调用新的send_partner_reports"""
        return self.send_partner_reports(pub_summary, feishu_upload_result, report_date)
    
    # 保持向后兼容性的方法别名
    def _send_single_pub_email(self, pub_name, email_data, file_paths, receivers, feishu_info, report_date=None):
        """向后兼容性方法，调用新的_send_single_partner_email"""
        return self._send_single_partner_email(pub_name, email_data, file_paths, receivers, feishu_info, report_date)
    
    # 保持向后兼容性的方法别名
    def _prepare_pub_email_data(self, pub_name, pub_data, report_date=None):
        """向后兼容性方法，调用新的_prepare_partner_email_data"""
        return self._prepare_partner_email_data(pub_name, pub_data, report_date)
    
    # 保持向后兼容性的方法别名
    def _get_pub_feishu_info(self, pub_name, feishu_upload_result):
        """向后兼容性方法，调用新的_get_partner_feishu_info"""
        return self._get_partner_feishu_info(pub_name, feishu_upload_result)
    
    def _create_pub_email_message(self, pub_name, email_data, file_paths, receivers, feishu_info, report_date=None, cc_email=None):
        """向后兼容性方法，调用新的_create_partner_email_message"""
        return self._create_partner_email_message(pub_name, email_data, file_paths, receivers, feishu_info, report_date, cc_email)

    def _is_bytec_partner(self, partner_name):
        """检查是否为ByteC Partner"""
        return config.is_bytec_partner(partner_name)

    def _generate_bytec_email_body(self, partner_name, email_data, feishu_info):
        """生成ByteC专用邮件正文"""
        report_date = email_data.get('report_date', datetime.now().strftime("%Y-%m-%d"))
        completion_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 从email_data获取基本信息
        start_date = email_data.get('start_date', report_date)
        end_date = email_data.get('end_date', report_date)
        main_file = email_data.get('main_file', f'ByteC_ConversionReport_{report_date}.xlsx')
        
        # 加载ByteC专用HTML模板
        template = self._load_html_template('bytec_email_template.html')
        if not template:
            # 如果模板加载失败，使用备用的简单HTML
            return self._generate_fallback_email_body(partner_name, email_data, feishu_info)
        
        # 从Excel文件计算ByteC三个维度的数据
        file_path = email_data.get('file_path')
        if not file_path or not os.path.exists(file_path):
            print_step("ByteC邮件", f"⚠️ 文件不存在: {file_path}")
            # 使用默认值
            bytec_data = self._get_default_bytec_data()
        else:
            bytec_data = self._calculate_bytec_summary_from_excel(file_path)
        
        # 准备飞书链接部分
        feishu_section = ""
        if feishu_info and self.include_feishu_links:
            if feishu_info.get('success') and feishu_info.get('uploaded_files'):
                feishu_template = self._load_html_template('feishu_section.html')
                if feishu_template:
                    feishu_links = ""
                    for file_info in feishu_info['uploaded_files']:
                        filename = file_info.get('filename', '')
                        
                        # 优先使用飞书返回的URL
                        feishu_url = file_info.get('url')
                        if not feishu_url:
                            # 如果没有url字段，尝试用file_token构造
                            file_token = file_info.get('file_token') or file_info.get('file_id')
                            if file_token:
                                feishu_url = f"https://bytedance.feishu.cn/sheets/{file_token}"
                        
                        if feishu_url:
                            feishu_links += f'<li><a href="{feishu_url}" target="_blank" style="color: #007bff; text-decoration: none; font-weight: bold;">{filename}</a></li>'
                        else:
                            feishu_links += f'<li>{filename} (已上传)</li>'
                    
                    feishu_section = feishu_template.replace('{{feishu_links}}', feishu_links)
        
        # 替换模板中的占位符
        body = template.replace('{{start_date}}', start_date)
        body = body.replace('{{end_date}}', end_date)
        body = body.replace('{{main_file}}', main_file)
        body = body.replace('{{completion_time}}', completion_time)
        body = body.replace('{{feishu_section}}', feishu_section)
        
        # ByteC Company Level Summary
        body = body.replace('{{company_total_conversion}}', str(bytec_data['company']['total_conversion']))
        body = body.replace('{{company_total_sales}}', bytec_data['company']['total_sales'])
        body = body.replace('{{company_total_earning}}', bytec_data['company']['total_earning'])
        body = body.replace('{{company_total_adv_commission}}', bytec_data['company']['total_adv_commission'])
        body = body.replace('{{company_total_pub_commission}}', bytec_data['company']['total_pub_commission'])
        body = body.replace('{{company_total_bytec_commission}}', bytec_data['company']['total_bytec_commission'])
        body = body.replace('{{company_bytec_roi}}', bytec_data['company']['bytec_roi'])
        body = body.replace('{{company_roi_class}}', bytec_data['company']['roi_class'])
        
        # Partner + Source Level Summary
        partner_source_rows = self._generate_partner_source_summary_rows(bytec_data['partner_source'])
        body = body.replace('{{partner_source_summary_rows}}', partner_source_rows)
        
        # Offer Level Summary
        offer_rows = self._generate_offer_summary_rows(bytec_data['offer'])
        body = body.replace('{{offer_summary_rows}}', offer_rows)
        
        return body

    def _calculate_bytec_summary_from_excel(self, file_path):
        """从Excel文件计算ByteC三个维度的汇总数据"""
        try:
            # 读取Excel文件的所有sheet
            excel_file = pd.ExcelFile(file_path)
            all_data = []
            
            # 合并所有sheet的数据
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                if not df.empty:
                    all_data.append(df)
            
            if not all_data:
                return self._get_default_bytec_data()
            
            # 合并所有数据
            combined_df = pd.concat(all_data, ignore_index=True)
            
            # 计算Company Level Summary
            company_summary = self._calculate_company_level_summary(combined_df)
            
            # 计算Partner + Source Level Summary (按优先级排序)
            partner_source_summary = self._calculate_partner_source_summary(combined_df)
            
            # 计算Offer Level Summary (按优先级排序)
            offer_summary = self._calculate_offer_level_summary(combined_df)
            
            return {
                'company': company_summary,
                'partner_source': partner_source_summary,
                'offer': offer_summary
            }
            
        except Exception as e:
            print_step("ByteC数据计算", f"❌ 计算失败: {str(e)}")
            return self._get_default_bytec_data()

    def _calculate_company_level_summary(self, df):
        """计算公司级别汇总"""
        try:
            # 找到转换数量列 - 关键修复！
            conversion_column = None
            for col in ['Conversions', 'conversions', 'conversion', 'Total Conversion']:
                if col in df.columns:
                    conversion_column = col
                    break
            
            # 基本统计：优先使用Conversions列求和，否则用行数
            if conversion_column:
                total_conversion = df[conversion_column].sum()
            else:
                total_conversion = len(df)
            
            # 金额计算 - 支持多种列名格式
            sales_amount_column = None
            for col in ['Sales Amount', 'sale_amount', 'sales_amount']:
                if col in df.columns:
                    sales_amount_column = col
                    break
            
            total_sales = 0.0
            if sales_amount_column:
                total_sales = df[sales_amount_column].sum()
            
            # Estimated Earning计算
            earning_column = None
            for col in ['Estimated Earning', 'estimated_earning', 'earning']:
                if col in df.columns:
                    earning_column = col
                    break
            
            total_earning = 0.0
            if earning_column:
                total_earning = df[earning_column].sum()
            
            # 佣金计算
            adv_commission_total = 0.0
            pub_commission_total = 0.0
            bytec_commission_total = 0.0
            
            # Adv Commission Rate计算
            adv_commission_column = None
            for col in ['Adv Commission Rate', 'adv_commission_rate']:
                if col in df.columns:
                    adv_commission_column = col
                    break
            
            if adv_commission_column and sales_amount_column:
                # Adv Commission = Sales Amount * Adv Commission Rate
                df_adv = df.copy()
                df_adv[adv_commission_column] = pd.to_numeric(df_adv[adv_commission_column], errors='coerce').fillna(0)
                df_adv[sales_amount_column] = pd.to_numeric(df_adv[sales_amount_column], errors='coerce').fillna(0)
                adv_commission_total = (df_adv[sales_amount_column] * df_adv[adv_commission_column]).sum()
            
            # Pub Commission Rate计算
            pub_commission_column = None
            for col in ['Pub Commission Rate', 'pub_commission_rate']:
                if col in df.columns:
                    pub_commission_column = col
                    break
            
            if pub_commission_column and sales_amount_column:
                # Pub Commission = Sales Amount * Pub Commission Rate
                df_pub = df.copy()
                df_pub[pub_commission_column] = pd.to_numeric(df_pub[pub_commission_column], errors='coerce').fillna(0)
                df_pub[sales_amount_column] = pd.to_numeric(df_pub[sales_amount_column], errors='coerce').fillna(0)
                pub_commission_total = (df_pub[sales_amount_column] * df_pub[pub_commission_column]).sum()
            
            # ByteC Commission = Adv Commission - Pub Commission
            bytec_commission_total = adv_commission_total - pub_commission_total
            
            # ByteC ROI计算 = ByteC Commission / Estimated Earning * 100%
            bytec_roi = 0.0
            roi_class = "amount"  # 默认绿色
            if total_earning > 0:
                bytec_roi = (bytec_commission_total / total_earning) * 100
                if bytec_roi < 0:
                    roi_class = "negative-roi"  # 负值用红色
            
            return {
                'total_conversion': int(total_conversion),  # 确保是整数
                'total_sales': f"${total_sales:,.2f}",
                'total_earning': f"${total_earning:,.2f}",
                'total_adv_commission': f"${adv_commission_total:,.2f}",
                'total_pub_commission': f"${pub_commission_total:,.2f}",
                'total_bytec_commission': f"${bytec_commission_total:,.2f}",
                'bytec_roi': f"{bytec_roi:.2f}%",
                'roi_class': roi_class
            }
            
        except Exception as e:
            print_step("公司汇总计算", f"❌ 失败: {str(e)}")
            return {
                'total_conversion': 0,
                'total_sales': "$0.00",
                'total_earning': "$0.00",
                'total_adv_commission': "$0.00",
                'total_pub_commission': "$0.00",
                'total_bytec_commission': "$0.00",
                'bytec_roi': "0.00%",
                'roi_class': "amount"
            }

    def _calculate_partner_source_summary(self, df):
        """计算Partner + Source维度汇总（按优先级排序）"""
        try:
            # 找到Partner和Source列
            partner_column = None
            source_column = None
            
            for col in ['Partner', 'partner', 'Partner Name']:
                if col in df.columns:
                    partner_column = col
                    break
            
            for col in ['Source', 'source', 'aff_sub1', 'Source Name']:
                if col in df.columns:
                    source_column = col
                    break
            
            if not partner_column or not source_column:
                return []
            
            # 找到转换数量列 - 关键修复！
            conversion_column = None
            for col in ['Conversions', 'conversions', 'conversion', 'Total Conversion']:
                if col in df.columns:
                    conversion_column = col
                    break
            
            # 找到金额列
            sales_amount_column = None
            earning_column = None
            
            for col in ['Sales Amount', 'sale_amount', 'sales_amount']:
                if col in df.columns:
                    sales_amount_column = col
                    break
                    
            for col in ['Estimated Earning', 'estimated_earning', 'earning']:
                if col in df.columns:
                    earning_column = col
                    break
            
            # 按Partner + Source分组统计
            agg_dict = {}
            if sales_amount_column:
                agg_dict[sales_amount_column] = 'sum'
            if earning_column:
                agg_dict[earning_column] = 'sum'
            # 关键修复：如果有Conversions列，则求和；否则用行数
            if conversion_column:
                agg_dict[conversion_column] = 'sum'
            
            grouped = df.groupby([partner_column, source_column]).agg(agg_dict).reset_index()
            
            # 计算转换数量：优先使用Conversions列，否则用行数
            if conversion_column:
                # 直接使用Conversions列的汇总值
                grouped['conversion_count'] = grouped[conversion_column]
            else:
                # 回退到计算分组行数（兼容性）
                group_counts = df.groupby([partner_column, source_column]).size().reset_index(name='conversion_count')
                grouped = grouped.merge(group_counts, on=[partner_column, source_column])
            
            # 重命名列
            column_mapping = {
                partner_column: partner_column,
                source_column: source_column
            }
            if sales_amount_column:
                column_mapping[sales_amount_column] = 'total_sales'
            if earning_column:
                column_mapping[earning_column] = 'total_earning'
                
            grouped = grouped.rename(columns=column_mapping)
            
            # 创建Partner + Source组合
            grouped['partner_source'] = grouped[partner_column] + "+" + grouped[source_column]
            
            # 确保missing列存在默认值
            if 'total_sales' not in grouped.columns:
                grouped['total_sales'] = 0.0
            if 'total_earning' not in grouped.columns:
                grouped['total_earning'] = 0.0
            
            # 按Estimated Earning降序排序（优先级）
            grouped = grouped.sort_values('total_earning', ascending=False)
            
            # 转换为列表格式
            summary_list = []
            for idx, row in grouped.iterrows():
                summary_list.append({
                    'partner_source': row['partner_source'],
                    'conversion': int(row['conversion_count']),
                    'sales_amount': f"${row['total_sales']:,.2f}",
                    'estimated_earning': f"${row['total_earning']:,.2f}"
                })
            
            return summary_list
            
        except Exception as e:
            print_step("Partner+Source汇总", f"❌ 失败: {str(e)}")
            return []

    def _calculate_offer_level_summary(self, df):
        """计算Offer维度汇总（按优先级排序）"""
        try:
            # 找到Offer列
            offer_column = None
            
            for col in ['Offer Name', 'offer_name', 'offer', 'Offer']:
                if col in df.columns:
                    offer_column = col
                    break
            
            if not offer_column:
                return []
            
            # 找到转换数量列 - 关键修复！
            conversion_column = None
            for col in ['Conversions', 'conversions', 'conversion', 'Total Conversion']:
                if col in df.columns:
                    conversion_column = col
                    break
            
            # 找到金额列
            sales_amount_column = None
            earning_column = None
            
            for col in ['Sales Amount', 'sale_amount', 'sales_amount']:
                if col in df.columns:
                    sales_amount_column = col
                    break
                    
            for col in ['Estimated Earning', 'estimated_earning', 'earning']:
                if col in df.columns:
                    earning_column = col
                    break
            
            # 过滤掉TOTAL行（这是Excel中的汇总行，不是真实的Offer）
            df_filtered = df[df[offer_column] != 'TOTAL'].copy()
            
            # 按Offer分组统计
            agg_dict = {}
            if sales_amount_column:
                agg_dict[sales_amount_column] = 'sum'
            if earning_column:
                agg_dict[earning_column] = 'sum'
            # 关键修复：如果有Conversions列，则求和；否则用行数
            if conversion_column:
                agg_dict[conversion_column] = 'sum'
            
            grouped = df_filtered.groupby(offer_column).agg(agg_dict).reset_index()
            
            # 计算转换数量：优先使用Conversions列，否则用行数
            if conversion_column:
                # 直接使用Conversions列的汇总值
                grouped['conversion_count'] = grouped[conversion_column]
            else:
                # 回退到计算分组行数（兼容性）
                group_counts = df_filtered.groupby(offer_column).size().reset_index(name='conversion_count')
                grouped = grouped.merge(group_counts, on=offer_column)
            
            # 重命名列
            column_mapping = {
                offer_column: offer_column
            }
            if sales_amount_column:
                column_mapping[sales_amount_column] = 'total_sales'
            if earning_column:
                column_mapping[earning_column] = 'total_earning'
                
            grouped = grouped.rename(columns=column_mapping)
            
            # 确保missing列存在默认值
            if 'total_sales' not in grouped.columns:
                grouped['total_sales'] = 0.0
            if 'total_earning' not in grouped.columns:
                grouped['total_earning'] = 0.0
            
            # 按Estimated Earning降序排序（优先级）
            grouped = grouped.sort_values('total_earning', ascending=False)
            
            # 转换为列表格式
            summary_list = []
            for idx, row in grouped.iterrows():
                summary_list.append({
                    'offer_name': row[offer_column],
                    'conversion': int(row['conversion_count']),
                    'sales_amount': f"${row['total_sales']:,.2f}",
                    'estimated_earning': f"${row['total_earning']:,.2f}"
                })
            
            return summary_list
            
        except Exception as e:
            print_step("Offer汇总", f"❌ 失败: {str(e)}")
            return []

    def _generate_partner_source_summary_rows(self, partner_source_data):
        """生成Partner + Source汇总表格行"""
        if not partner_source_data:
            return "<tr><td colspan='5'>暂无数据</td></tr>"
        
        rows = []
        for idx, item in enumerate(partner_source_data, 1):
            row = f"""
            <tr>
                <td><strong>{idx}</strong></td>
                <td><strong>{item['partner_source']}</strong></td>
                <td>{item['conversion']}</td>
                <td><span class="amount">{item['sales_amount']}</span></td>
                <td><span class="amount">{item['estimated_earning']}</span></td>
            </tr>
            """
            rows.append(row)
        
        return "".join(rows)

    def _generate_offer_summary_rows(self, offer_data):
        """生成Offer汇总表格行"""
        if not offer_data:
            return "<tr><td colspan='5'>暂无数据</td></tr>"
        
        rows = []
        for idx, item in enumerate(offer_data, 1):
            row = f"""
            <tr>
                <td><strong>{idx}</strong></td>
                <td><strong>{item['offer_name']}</strong></td>
                <td>{item['conversion']}</td>
                <td><span class="amount">{item['sales_amount']}</span></td>
                <td><span class="amount">{item['estimated_earning']}</span></td>
            </tr>
            """
            rows.append(row)
        
        return "".join(rows)

    def _get_default_bytec_data(self):
        """获取默认的ByteC数据（当无法读取Excel时使用）"""
        return {
            'company': {
                'total_conversion': 0,
                'total_sales': "$0.00",
                'total_earning': "$0.00",
                'total_adv_commission': "$0.00",
                'total_pub_commission': "$0.00",
                'total_bytec_commission': "$0.00",
                'bytec_roi': "0.00%",
                'roi_class': "amount"
            },
            'partner_source': [],
            'offer': []
        }

# 便捷函数
def send_report_email(report_data, file_paths=None, feishu_upload_result=None):
    """
    便捷的邮件发送函数
    
    Args:
        report_data: 报告数据字典
        file_paths: Excel文件路径列表
        feishu_upload_result: 飞书上传结果
    
    Returns:
        dict: 发送结果
    """
    sender = EmailSender()
    return sender.send_report_email(report_data, file_paths, feishu_upload_result)

def test_email_connection():
    """
    测试邮件连接的便捷函数
    
    Returns:
        bool: 连接是否成功
    """
    sender = EmailSender()
    return sender.test_connection() 