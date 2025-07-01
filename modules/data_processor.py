#!/usr/bin/env python3
"""
数据清洗与调整模块
负责处理从Involve Asia获取的原始数据，包括数据清洗、格式化、分类导出等功能
"""

import pandas as pd
import os
from datetime import datetime
from utils.logger import print_step
import config
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

class DataProcessor:
    """数据处理器类"""
    
    def __init__(self):
        self.original_data = None
        self.processed_data = None
        self.total_sale_amount = 0
        self.pub_summary = {}
        self.report_date = None
    
    def process_data(self, data_source, output_dir=None, report_date=None, start_date=None, end_date=None):
        """
        完整的数据处理流程
        
        Args:
            data_source: 数据源（DataFrame、Excel文件路径或JSON数据）
            output_dir: 输出目录，默认使用config.OUTPUT_DIR
            report_date: 报告日期，用于文件名，默认使用当前日期（向后兼容）
            start_date: 开始日期，用于文件名生成
            end_date: 结束日期，用于文件名生成
        
        Returns:
            dict: 处理结果摘要
        """
        print_step("数据处理开始", "开始执行完整的数据清洗与调整流程")
        
        if output_dir is None:
            output_dir = config.OUTPUT_DIR
        
        # 设置报告日期范围
        if report_date:
            self.report_date = report_date
        if start_date:
            self.start_date = start_date
        if end_date:
            self.end_date = end_date
        
        # 步骤1: 加载数据
        self._load_data(data_source)
        
        # 步骤2: 数据清洗
        self._clean_data()
        
        # 步骤3: 格式化金额并统计
        self._format_and_calculate_amounts()
        
        # 步骤4: 调整金额（Mockup）
        self._apply_mockup_adjustment()
        
        # 步骤5: 按Pub分类导出
        pub_files = self._export_by_pub(output_dir)
        
        # 步骤6: 生成处理摘要
        result = self._generate_summary(pub_files, output_dir)
        
        print_step("数据处理完成", "所有数据处理步骤执行完毕")
        return result
    
    def _load_data(self, data_source):
        """加载数据源"""
        print_step("数据加载", "正在加载原始数据...")
        
        if isinstance(data_source, pd.DataFrame):
            self.original_data = data_source.copy()
        elif isinstance(data_source, str) and data_source.endswith('.xlsx'):
            self.original_data = pd.read_excel(data_source)
        elif isinstance(data_source, str) and data_source.endswith('.json'):
            # 处理JSON格式数据
            import json
            with open(data_source, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            if 'data' in json_data and 'data' in json_data['data']:
                conversion_records = json_data['data']['data']
                self.original_data = pd.DataFrame(conversion_records)
            elif 'data' in json_data and 'conversions' in json_data['data']:
                # 支持多API合并格式
                conversion_records = json_data['data']['conversions']
                self.original_data = pd.DataFrame(conversion_records)
            else:
                raise ValueError("JSON数据格式不正确")
        elif isinstance(data_source, dict):
            # 直接处理字典格式的JSON数据
            if 'data' in data_source and 'data' in data_source['data']:
                conversion_records = data_source['data']['data']
                self.original_data = pd.DataFrame(conversion_records)
            elif 'data' in data_source and 'conversions' in data_source['data']:
                # 支持多API合并格式
                conversion_records = data_source['data']['conversions']
                self.original_data = pd.DataFrame(conversion_records)
            else:
                raise ValueError("JSON数据格式不正确")
        else:
            raise ValueError(f"不支持的数据源格式: {type(data_source)}")
        
        print_step("数据加载完成", f"成功加载 {len(self.original_data)} 条记录，{len(self.original_data.columns)} 个字段")
        
        # 复制到处理数据
        self.processed_data = self.original_data.copy()
    
    def _clean_data(self):
        """数据清洗 - 移除不需要的栏位"""
        print_step("数据清洗", f"正在移除不需要的栏位: {config.REMOVE_COLUMNS}")
        
        # 检查哪些栏位实际存在
        existing_columns = [col for col in config.REMOVE_COLUMNS if col in self.processed_data.columns]
        missing_columns = [col for col in config.REMOVE_COLUMNS if col not in self.processed_data.columns]
        
        if existing_columns:
            self.processed_data = self.processed_data.drop(columns=existing_columns)
            print_step("栏位移除", f"成功移除栏位: {existing_columns}")
        
        if missing_columns:
            print_step("栏位警告", f"以下栏位不存在，跳过移除: {missing_columns}")
        
        print_step("清洗完成", f"清洗后剩余 {len(self.processed_data.columns)} 个字段")
    
    def _format_and_calculate_amounts(self):
        """格式化金额并计算总值"""
        print_step("金额处理", "正在格式化sale_amount栏位并计算总值...")
        
        if 'sale_amount' not in self.processed_data.columns:
            print_step("金额处理警告", "sale_amount栏位不存在，跳过金额处理")
            return
        
        # 转换为数值类型并处理异常值
        self.processed_data['sale_amount'] = pd.to_numeric(
            self.processed_data['sale_amount'], 
            errors='coerce'
        ).fillna(0)
        
        # 格式化为两位小数
        self.processed_data['sale_amount'] = self.processed_data['sale_amount'].round(2)
        
        # 计算总值
        self.total_sale_amount = self.processed_data['sale_amount'].sum()
        
        print_step("金额统计", f"sale_amount总值: ${self.total_sale_amount:,.2f} USD")
        print_step("金额格式化完成", f"所有金额已格式化为美金格式（小数点后两位）")
    
    def _apply_mockup_adjustment(self):
        """应用Mockup调整倍数"""
        print_step("金额调整", f"正在应用Mockup调整倍数: {config.MOCKUP_MULTIPLIER * 100}%")
        
        if 'sale_amount' not in self.processed_data.columns:
            print_step("金额调整警告", "sale_amount栏位不存在，跳过金额调整")
            return
        
        # 保存调整前的总值
        original_total = self.processed_data['sale_amount'].sum()
        
        # 应用调整倍数
        self.processed_data['sale_amount'] = (
            self.processed_data['sale_amount'] * config.MOCKUP_MULTIPLIER
        ).round(2)
        
        # 计算调整后的总值
        adjusted_total = self.processed_data['sale_amount'].sum()
        
        print_step("金额调整完成", f"调整前总值: ${original_total:,.2f} → 调整后总值: ${adjusted_total:,.2f}")
        
        # 更新总值
        self.total_sale_amount = adjusted_total
    
    def _export_by_pub(self, output_dir):
        """按Partner分类导出Excel文件，每个Partner包含多个Sources作为Sheets"""
        print_step("Partner分类导出", "正在按Partner分类导出，每个Partner包含多个Sources作为Sheets...")
        
        if 'aff_sub1' not in self.processed_data.columns:
            print_step("分类导出警告", "aff_sub1 (Source) 栏位不存在，跳过分类导出")
            return []
        
        # 获取所有唯一的Source值（aff_sub1）
        unique_sources = self.processed_data['aff_sub1'].dropna().unique()
        print_step("Source统计", f"发现 {len(unique_sources)} 个不同的Source: {list(unique_sources)}")
        
        # 按Partner分组Sources
        partner_sources_map = {}
        for source in unique_sources:
            partner = config.match_source_to_partner(source)
            if partner not in partner_sources_map:
                partner_sources_map[partner] = []
            partner_sources_map[partner].append(source)
        
        print_step("Partner映射", f"映射结果: {partner_sources_map}")
        
        partner_files = []
        # 确定文件名的日期范围
        if hasattr(self, 'start_date') and hasattr(self, 'end_date') and self.start_date and self.end_date:
            # 使用真实的日期范围
            start_date = self.start_date
            end_date = self.end_date
        elif hasattr(self, 'report_date') and self.report_date:
            # 向后兼容：使用单个报告日期
            start_date = self.report_date
            end_date = self.report_date
        else:
            # 默认使用当前日期
            current_date = datetime.now().strftime("%Y-%m-%d")
            start_date = current_date
            end_date = current_date
        
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        
        for partner, sources_list in partner_sources_map.items():
            # 检查是否在目标Partner列表中
            if not config.is_partner_enabled(partner):
                print_step("Partner跳过", f"跳过Partner '{partner}' (不在处理范围内)")
                continue
                
            try:
                # 生成Partner文件名
                filename = config.get_partner_filename(partner, start_date, end_date)
                filepath = os.path.join(output_dir, filename)
                
                # 创建Excel工作簿，包含多个Sources作为Sheets
                self._create_partner_excel_with_sources(partner, sources_list, filepath)
                
                # 统计Partner总信息
                partner_data = self.processed_data[self.processed_data['aff_sub1'].isin(sources_list)]
                partner_total = partner_data['sale_amount'].sum() if 'sale_amount' in partner_data.columns else 0
                
                self.pub_summary[partner] = {
                    'records': len(partner_data),
                    'total_amount': partner_total,
                    'amount_formatted': f"${partner_total:,.2f}",
                    'filename': filename,
                    'sources': sources_list,
                    'sources_count': len(sources_list)
                }
                
                partner_files.append(filepath)
                
                print_step("Partner导出", f"Partner '{partner}': {len(sources_list)} 个Sources, {len(partner_data)} 条记录，总金额 ${partner_total:,.2f} → {filename}")
                
            except Exception as e:
                print_step("Partner导出错误", f"❌ Partner '{partner}' 导出失败: {str(e)}")
                continue
        
        print_step("分类导出完成", f"成功生成 {len(partner_files)} 个Partner分类文件")
        return partner_files
    
    def _create_partner_excel_with_sources(self, partner, sources_list, filepath):
        """
        创建Partner Excel文件，包含多个Sources作为不同的Sheets
        
        Args:
            partner: Partner名称
            sources_list: 该Partner下的Sources列表
            filepath: 输出文件路径
        """
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # 创建工作簿
        wb = Workbook()
        # 删除默认的工作表
        wb.remove(wb.active)
        
        # 为每个Source创建一个Sheet
        for source in sources_list:
            # 过滤该Source的数据
            source_data = self.processed_data[self.processed_data['aff_sub1'] == source].copy()
            
            if len(source_data) == 0:
                print_step("Sheet创建", f"⚠️ Source '{source}' 没有数据，跳过创建Sheet")
                continue
            
            # 创建工作表，使用Source名称作为Sheet名（清理特殊字符并限制长度）
            safe_sheet_name = self._clean_sheet_name(str(source))
            ws = wb.create_sheet(title=safe_sheet_name)
            
            # 写入数据（包含标题行），清理特殊字符
            for r in dataframe_to_rows(source_data, index=False, header=True):
                # 清理行中的特殊字符
                cleaned_row = self._clean_row_data(r)
                ws.append(cleaned_row)
            
            # 查找sale_amount列的索引并应用货币格式
            if 'sale_amount' in source_data.columns:
                sale_amount_col = source_data.columns.get_loc('sale_amount') + 1  # Excel列索引从1开始
                
                # 应用货币格式到sale_amount列（跳过标题行）
                for row in range(2, len(source_data) + 2):  # 从第2行开始（第1行是标题）
                    cell = ws.cell(row=row, column=sale_amount_col)
                    cell.number_format = '"$"#,##0.00'
            
            print_step("Sheet创建", f"✅ 已创建Sheet '{safe_sheet_name}' ({len(source_data)} 条记录)")
        
        # 检查是否有任何Sheet被创建
        if len(wb.worksheets) == 0:
            # 如果没有Sheet，创建一个空的Sheet
            ws = wb.create_sheet(title="No_Data")
            ws.append(["Partner", "Message"])
            ws.append([partner, "No data available"])
            print_step("Sheet创建", f"⚠️ Partner '{partner}' 没有任何数据，创建空Sheet")
        
        # 保存文件
        wb.save(filepath)
        print_step("Excel保存", f"✅ Partner Excel文件已保存: {filepath} (包含 {len(wb.worksheets)} 个Sheets)")
    
    def _clean_sheet_name(self, name):
        """
        清理Excel工作表名称，移除不支持的字符
        
        Excel工作表名称限制：
        - 不能超过31个字符
        - 不能包含: [ ] : * ? / \\ '
        - 不能为空或只包含空格
        - 不能以单引号开头或结尾
        
        Args:
            name: 原始名称
            
        Returns:
            str: 清理后的安全名称
        """
        import re
        
        if not name or not str(name).strip():
            return "Unknown"
        
        # 转换为字符串并去除前后空格
        clean_name = str(name).strip()
        
        # 移除或替换Excel不支持的字符
        # 替换路径分隔符
        clean_name = clean_name.replace('/', '_').replace('\\', '_')
        # 替换Excel特殊字符
        clean_name = clean_name.replace('[', '(').replace(']', ')')
        clean_name = clean_name.replace(':', '-').replace('*', '_')
        clean_name = clean_name.replace('?', '_').replace('\'', '')
        
        # 移除其他可能有问题的Unicode字符，保留基本字母、数字、空格和常见符号
        # 使用正则表达式保留安全字符
        clean_name = re.sub(r'[^\w\s\-\(\)\_\.]', '_', clean_name)
        
        # 移除多余的空格和下划线
        clean_name = re.sub(r'\s+', ' ', clean_name)  # 多个空格合并为一个
        clean_name = re.sub(r'_+', '_', clean_name)   # 多个下划线合并为一个
        clean_name = clean_name.strip('_').strip()    # 去除开头结尾的下划线和空格
        
        # 确保不以单引号开头或结尾
        clean_name = clean_name.strip('\'')
        
        # 限制长度为31个字符
        if len(clean_name) > 31:
            clean_name = clean_name[:28] + "..."
        
        # 如果清理后为空，使用默认名称
        if not clean_name:
            clean_name = "Unknown"
        
        return clean_name
    
    def _clean_row_data(self, row):
        """
        清理行数据中的特殊字符，确保Excel兼容性
        
        Args:
            row: 数据行（列表或元组）
            
        Returns:
            list: 清理后的数据行
        """
        import re
        
        cleaned_row = []
        for cell in row:
            if cell is None:
                cleaned_row.append(None)
            elif isinstance(cell, (int, float)):
                # 数字类型直接保留
                cleaned_row.append(cell)
            else:
                # 字符串类型需要清理
                cell_str = str(cell)
                
                # 移除可能导致Excel问题的控制字符和特殊Unicode字符
                # 保留基本的ASCII字符、常见Unicode字符
                cleaned_str = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', cell_str)
                
                # 移除可能有问题的Unicode字符（保留基本字母、数字、常见符号）
                # 这个正则表达式比较宽松，保留大部分字符但移除控制字符
                cleaned_str = re.sub(r'[^\x20-\x7E\u00A0-\u024F\u1E00-\u1EFF\u2000-\u206F\u20A0-\u20CF\u2100-\u214F]', '_', cleaned_str)
                
                # 移除开头的特殊字符（如不可见字符）
                cleaned_str = cleaned_str.strip()
                
                cleaned_row.append(cleaned_str)
        
        return cleaned_row
    
    def _generate_summary(self, pub_files, output_dir):
        """生成处理结果摘要"""
        print_step("生成摘要", "正在生成数据处理结果摘要...")
        
        summary = {
            'success': True,
            'original_records': len(self.original_data) if self.original_data is not None else 0,
            'processed_records': len(self.processed_data) if self.processed_data is not None else 0,
            'total_sale_amount': self.total_sale_amount,
            'mockup_multiplier': config.MOCKUP_MULTIPLIER,
            'removed_columns': config.REMOVE_COLUMNS,
            'partner_count': len(self.pub_summary),
            'pub_count': len(self.pub_summary),
            'pub_files': pub_files,
            'partner_summary': self.pub_summary,
            'pub_summary': self.pub_summary,
            'output_directory': output_dir
        }
        
        return summary
    
    def print_detailed_summary(self, summary):
        """打印详细的处理摘要"""
        print_step("处理摘要", "数据处理完成，详细结果如下:")
        
        print(f"📊 数据处理统计:")
        print(f"   - 原始记录数: {summary['original_records']:,}")
        print(f"   - 处理后记录数: {summary['processed_records']:,}")
        print(f"   - 移除栏位: {', '.join(summary['removed_columns'])}")
        print(f"   - Mockup调整倍数: {summary['mockup_multiplier'] * 100}%")
        
        print(f"💰 金额统计:")
        print(f"   - 调整后总金额: ${summary['total_sale_amount']:,.2f} USD")
        
        print(f"📂 Pub分类导出:")
        print(f"   - 不同Pub数量: {summary['pub_count']}")
        print(f"   - 生成文件数量: {len(summary['pub_files'])}")
        print(f"   - 输出目录: {summary['output_directory']}")
        
        if summary['pub_summary']:
            print(f"📋 各Pub详细信息:")
            for pub, info in summary['pub_summary'].items():
                print(f"   - {pub}: {info['records']} 条记录, ${info['total_amount']:,.2f}, 文件: {info['filename']}")

# 便捷函数
def process_conversion_data(data_source, output_dir=None):
    """
    便捷的数据处理函数
    
    Args:
        data_source: 数据源
        output_dir: 输出目录
    
    Returns:
        dict: 处理结果摘要
    """
    processor = DataProcessor()
    summary = processor.process_data(data_source, output_dir)
    processor.print_detailed_summary(summary)
    return summary 